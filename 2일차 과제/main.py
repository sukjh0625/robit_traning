import openpyxl
from openpyxl.chart import LineChart, Reference

def main():
    print("=== OpenPyXL을 이용한 샤페론 주가 데이터 및 그래프 생성 ===")
    
    wb = openpyxl.Workbook()
    ws = wb.active      
    ws.title = "샤페론 2026년 7월 주가 변동" 
    
    # 데이터 입력
    ws.append(["날짜", "주가 변동률(%)", "거래량", "종가", "거래대금"])
    ws.append(["7/1", 6.88, 3073809, 543])
    ws.append(["7/2", -2.39, 2027860, 530])
    ws.append(["7/3", 13.96, 2374090, 604])
    ws.append(["7/6", -11.09, 1528981, 537])
    ws.append(["7/7", -2.04, 1893762, 526])
    ws.append(["7/8", -4.37, 997057, 503])
    ws.append(["7/9", 3.18, 775875, 519])
    ws.append(["7/10", 7.32, 1701094, 557])
    ws.append(["7/13", -6.1, 825717, 523])
    ws.append(["7/14", -2.29, 624211, 511])
    
    # 거래대금 자동 계산
    for row_num in range(2, 12):
        ws[f"E{row_num}"] = f"=C{row_num}*D{row_num}"
    
    # 평균 종가 계산
    ws["C13"] = "평균 종가"
    ws["D13"] = "=AVERAGE(D2:D11)"
    
    #차트 그리기
    chart = LineChart()
    chart.title = "샤페론 일별 종가 추이"   
    chart.style = 10                       # 텍스트 스타일
    chart.y_axis.title = "종가 (원)"        #y축
    chart.x_axis.title = "날짜"             #x축
    chart.legend = None                     
    
    
    chart.y_axis.scaling.min = 480          # 최저점 눈금
    chart.y_axis.scaling.max = 620          # 최고점 눈금
    chart.y_axis.majorUnit = 20             # 20원 단위로 눈금 표시
    
    
    data = Reference(ws, min_col=4, min_row=1, max_row=11) # 마우스 드래그 하는것처럼 지정된 범위를 그리라는 코드이다. y축을 나타내고 있다
    categories = Reference(ws, min_col=1, min_row=2, max_row=11) # 얘는 x축 즉 날짜 나타내는거다
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    
    ws.add_chart(chart, "G2") #g2위치에 차트 그리기
    
    
    output_filename = "shaperon_stock.xlsx" #엑셀 파일로 저장
    wb.save(output_filename)
    
    print(f"\n[완료]  '{output_filename}' 파일이 생성되었습니다!")

if __name__ == "__main__":
    main()
